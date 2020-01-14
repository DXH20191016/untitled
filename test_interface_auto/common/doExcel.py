from openpyxl import load_workbook

class doExcel():

    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
        self.wb=load_workbook(self.file_path)
        self.sh=self.wb[self.sheet_name]
        #获取当前sheet最大的行数
        self.row_max=self.sh.max_row
        # 获取当前sheet最大的列数
        self.column_max=self.sh.max_column

    def get_sheet(self):
        sheet_list=[]
        for i in range(2,self.row_max+1):
            sheet_row_list={}
            sheet_row_list["interface_name"]=self.sh.cell(i,1).value
            sheet_row_list["interface_url"]=self.sh.cell(i,2).value
            sheet_row_list["interface_method"]=self.sh.cell(i,3).value
            sheet_row_list["interface_body"]=self.sh.cell(i,4).value
            sheet_row_list["interface_expectedOutcome"]=self.sh.cell(i,5).value
            sheet_row_list["interface_actualResults"]=self.sh.cell(i,6).value
            sheet_list.append(sheet_row_list)
        return sheet_list

    def get_cell(self,row,column):
        return self.sh.cell(row,column).value

    def set_cell(self,row,column,cellvalue):
        self.sh.cell(row=row, column=column).value = cellvalue
        self.wb.save(self.file_path)

file_path="C:\\Users\\admin\\Desktop\\test_api.xlsx"
sheet_name="Sheet1"
do=doExcel(file_path,sheet_name)
do.set_cell(10,10,"")
print(do.get_cell(10,10))
